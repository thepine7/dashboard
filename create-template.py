#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 템플릿 생성 스크립트
차트가 포함된 엑셀 템플릿을 자동으로 생성합니다.
"""

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl 라이브러리가 설치되지 않았습니다.")
    print("다음 명령어로 설치하세요:")
    print("  pip install openpyxl")
    exit(1)

def create_excel_template():
    """차트가 포함된 엑셀 템플릿 생성"""
    
    print("📊 엑셀 템플릿 생성 시작...")
    
    # 워크북 생성
    wb = Workbook()
    
    # === 일간 데이터 시트 생성 ===
    ws_daily = wb.active
    ws_daily.title = "일간 데이터"
    print("✅ 일간 데이터 시트 생성")
    
    # 스타일 정의
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 생성 (1행)
    print("  - 헤더 생성 중... (1440개 시간 레이블)")
    
    # A1: 날짜명
    ws_daily['A1'] = "날짜명"
    ws_daily['A1'].fill = header_fill
    ws_daily['A1'].font = header_font
    ws_daily['A1'].alignment = center_align
    ws_daily['A1'].border = border
    
    # B1~CX1: 시간 레이블 (0:00 ~ 23:59, 1440개)
    for minute in range(1440):
        col = minute + 2  # B열부터 시작 (A=1, B=2)
        hour = minute // 60
        min_val = minute % 60
        time_label = f"{hour}:{min_val:02d}"
        
        cell = ws_daily.cell(row=1, column=col)
        cell.value = time_label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
        # 진행 상황 표시 (매 100개마다)
        if (minute + 1) % 100 == 0:
            print(f"    진행: {minute + 1}/1440 ({(minute + 1) / 1440 * 100:.1f}%)")
    
    print("  ✅ 헤더 생성 완료")
    
    # 샘플 데이터 행 생성 (2~10행, 최대 9일)
    print("  - 샘플 데이터 행 생성 중... (9일치)")
    sample_dates = [
        "2025-10-16", "2025-10-15", "2025-10-14",
        "2025-10-13", "2025-10-12", "2025-10-11",
        "2025-10-10", "2025-10-09", "2025-10-08"
    ]
    
    for row_idx, date in enumerate(sample_dates, start=2):
        # A열: 날짜명
        cell = ws_daily.cell(row=row_idx, column=1)
        cell.value = date
        cell.alignment = center_align
        cell.border = border
        
        # B열~CX열: 샘플 온도 데이터 (24.0~25.0 범위)
        for col in range(2, 1442):  # 1440개 + 2 (B열부터 시작)
            cell = ws_daily.cell(row=row_idx, column=col)
            # 샘플 데이터: 24.0 + 랜덤 변동
            import random
            cell.value = round(24.0 + random.uniform(0, 1.0), 1)
            cell.alignment = center_align
            cell.border = border
        
        print(f"    날짜 {date} 데이터 생성 완료 ({row_idx - 1}/9)")
    
    print("  ✅ 샘플 데이터 생성 완료")
    
    # 열 너비 조정
    print("  - 열 너비 조정 중...")
    ws_daily.column_dimensions['A'].width = 12  # 날짜명 열
    for col in range(2, 62):  # B~BJ열 (처음 60개 시간만)
        ws_daily.column_dimensions[get_column_letter(col)].width = 6
    print("  ✅ 열 너비 조정 완료")
    
    # 차트 생성
    print("  - 차트 생성 중...")
    chart = LineChart()
    chart.title = "일간 온도 추이"
    chart.style = 2
    chart.y_axis.title = "온도 (°C)"
    chart.x_axis.title = "시간"
    chart.height = 15  # 차트 높이
    chart.width = 25   # 차트 너비
    
    # 차트 데이터 추가 (처음 60개 시간만 - 1시간 간격)
    # X축: B1~BJ1 (시간 레이블)
    # Y축: B2~BJ10 (각 날짜별 데이터)
    print("    - 차트 시리즈 추가 중...")
    
    # 각 날짜별로 시리즈 생성 (최대 9개)
    for row_idx in range(2, 11):  # 2~10행
        # X축 데이터 (시간 레이블)
        cats = Reference(ws_daily, min_col=2, max_col=61, min_row=1, max_row=1)
        
        # Y축 데이터 (온도 값)
        data = Reference(ws_daily, min_col=2, max_col=61, min_row=row_idx, max_row=row_idx)
        
        # 시리즈 추가
        chart.add_data(data, titles_from_data=False)
        
        print(f"      시리즈 {row_idx - 1}: {ws_daily.cell(row=row_idx, column=1).value}")
    
    # X축 레이블 설정
    chart.set_categories(cats)
    
    # 차트 위치: A12 (데이터 테이블 아래)
    ws_daily.add_chart(chart, "A12")
    print("  ✅ 차트 생성 완료 (위치: A12)")
    
    # === 월간 데이터 시트 생성 ===
    print("\n✅ 월간 데이터 시트 생성")
    ws_monthly = wb.create_sheet("월간 데이터")
    
    # 헤더 생성 (1행)
    ws_monthly['A1'] = "날짜명"
    ws_monthly['A1'].fill = header_fill
    ws_monthly['A1'].font = header_font
    ws_monthly['A1'].alignment = center_align
    ws_monthly['A1'].border = border
    
    # B1~AF1: 일자 (1일~31일)
    for day in range(1, 32):
        col = day + 1  # B열부터 시작
        cell = ws_monthly.cell(row=1, column=col)
        cell.value = f"{day}일"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    print("  ✅ 월간 데이터 헤더 생성 완료")
    
    # 샘플 데이터 행 생성 (2~10행, 최대 9개월)
    sample_months = [
        "2025-09", "2025-08", "2025-07",
        "2025-06", "2025-05", "2025-04",
        "2025-03", "2025-02", "2025-01"
    ]
    
    for row_idx, month in enumerate(sample_months, start=2):
        # A열: 년월
        cell = ws_monthly.cell(row=row_idx, column=1)
        cell.value = month
        cell.alignment = center_align
        cell.border = border
        
        # B열~AF열: 샘플 온도 데이터
        for col in range(2, 33):  # 1~31일
            cell = ws_monthly.cell(row=row_idx, column=col)
            cell.value = round(24.0 + random.uniform(0, 1.0), 1)
            cell.alignment = center_align
            cell.border = border
    
    print("  ✅ 월간 데이터 샘플 생성 완료")
    
    # 열 너비 조정
    ws_monthly.column_dimensions['A'].width = 12
    for col in range(2, 33):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 6
    
    # === 파일 저장 ===
    output_path = "src/main/resources/excel-templates/chart-template.xlsx"
    print(f"\n💾 파일 저장 중: {output_path}")
    
    try:
        wb.save(output_path)
        print(f"✅ 엑셀 템플릿 생성 완료!")
        print(f"📁 파일 위치: {output_path}")
        print(f"📊 일간 데이터: 1440개 시간 x 9일 + 차트")
        print(f"📊 월간 데이터: 31일 x 9개월")
        print(f"\n🚀 다음 단계:")
        print(f"  1. 톰캣 재시작")
        print(f"  2. 엑셀 다운로드 테스트")
        print(f"  3. 차트 확인")
    except PermissionError:
        print(f"❌ 파일 저장 실패: 권한 오류")
        print(f"   파일이 열려있는지 확인하세요.")
    except Exception as e:
        print(f"❌ 파일 저장 실패: {e}")

if __name__ == "__main__":
    print("=" * 60)
    print("📊 엑셀 템플릿 자동 생성 스크립트")
    print("=" * 60)
    print()
    
    create_excel_template()
    
    print()
    print("=" * 60)
    print("✅ 완료!")
    print("=" * 60)

