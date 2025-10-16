#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 템플릿 생성 스크립트 (차트 없이 데이터 구조만)
사용자가 Excel에서 직접 차트를 추가할 수 있도록 데이터만 생성합니다.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl 라이브러리가 설치되지 않았습니다.")
    print("다음 명령어로 설치하세요:")
    print("  pip install openpyxl")
    exit(1)

def create_excel_template_simple():
    """차트 없이 데이터 구조만 있는 엑셀 템플릿 생성"""
    
    print("📊 엑셀 템플릿 생성 시작 (데이터 구조만)...")
    
    # 워크북 생성
    wb = Workbook()
    
    # === 일간 데이터 시트 생성 ===
    ws_daily = wb.active
    ws_daily.title = "일간 데이터"
    print("✅ 일간 데이터 시트 생성")
    
    # 스타일 정의
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 생성 (1행)
    print("  - 헤더 생성 중... (1440개 시간 레이블)")
    
    # A1: 날짜명
    ws_daily['A1'] = "날짜명"
    ws_daily['A1'].fill = header_fill
    ws_daily['A1'].font = header_font
    ws_daily['A1'].alignment = center_align
    ws_daily['A1'].border = border
    
    # B1~CX1: 시간 레이블 (0:00 ~ 23:59, 1440개)
    for minute in range(1440):
        col = minute + 2  # B열부터 시작
        hour = minute // 60
        min_val = minute % 60
        time_label = f"{hour}:{min_val:02d}"
        
        cell = ws_daily.cell(row=1, column=col)
        cell.value = time_label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
        if (minute + 1) % 100 == 0:
            print(f"    진행: {minute + 1}/1440 ({(minute + 1) / 1440 * 100:.1f}%)")
    
    print("  ✅ 헤더 생성 완료")
    
    # 샘플 데이터 행 생성 (2~10행, 최대 9일)
    print("  - 샘플 데이터 행 생성 중... (9일치)")
    sample_dates = [
        "2025-10-16", "2025-10-15", "2025-10-14",
        "2025-10-13", "2025-10-12", "2025-10-11",
        "2025-10-10", "2025-10-09", "2025-10-08"
    ]
    
    import random
    for row_idx, date in enumerate(sample_dates, start=2):
        # A열: 날짜명
        cell = ws_daily.cell(row=row_idx, column=1)
        cell.value = date
        cell.alignment = center_align
        cell.border = border
        
        # B열~CX열: 샘플 온도 데이터
        for col in range(2, 1442):
            cell = ws_daily.cell(row=row_idx, column=col)
            cell.value = round(24.0 + random.uniform(0, 1.0), 1)
            cell.alignment = center_align
            cell.border = border
        
        print(f"    날짜 {date} 데이터 생성 완료 ({row_idx - 1}/9)")
    
    print("  ✅ 샘플 데이터 생성 완료")
    
    # 열 너비 조정
    ws_daily.column_dimensions['A'].width = 12
    for col in range(2, 62):
        ws_daily.column_dimensions[get_column_letter(col)].width = 6
    
    # === 월간 데이터 시트 생성 ===
    print("\n✅ 월간 데이터 시트 생성")
    ws_monthly = wb.create_sheet("월간 데이터")
    
    # 헤더 생성
    ws_monthly['A1'] = "날짜명"
    ws_monthly['A1'].fill = header_fill
    ws_monthly['A1'].font = header_font
    ws_monthly['A1'].alignment = center_align
    ws_monthly['A1'].border = border
    
    for day in range(1, 32):
        col = day + 1
        cell = ws_monthly.cell(row=1, column=col)
        cell.value = f"{day}일"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 샘플 데이터
    sample_months = [
        "2025-09", "2025-08", "2025-07",
        "2025-06", "2025-05", "2025-04",
        "2025-03", "2025-02", "2025-01"
    ]
    
    for row_idx, month in enumerate(sample_months, start=2):
        cell = ws_monthly.cell(row=row_idx, column=1)
        cell.value = month
        cell.alignment = center_align
        cell.border = border
        
        for col in range(2, 33):
            cell = ws_monthly.cell(row=row_idx, column=col)
            cell.value = round(24.0 + random.uniform(0, 1.0), 1)
            cell.alignment = center_align
            cell.border = border
    
    ws_monthly.column_dimensions['A'].width = 12
    for col in range(2, 33):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 6
    
    # === 파일 저장 ===
    output_path = "chart-template-base.xlsx"
    print(f"\n💾 파일 저장 중: {output_path}")
    
    try:
        wb.save(output_path)
        print(f"✅ 기본 템플릿 생성 완료!")
        print(f"📁 파일 위치: {output_path}")
        print(f"\n📝 다음 단계:")
        print(f"  1. Excel에서 '{output_path}' 파일 열기")
        print(f"  2. 데이터 범위 선택 (A1:BJ10)")
        print(f"  3. 삽입 → 차트 → 꺾은선형 선택")
        print(f"  4. 차트를 A12 위치로 이동")
        print(f"  5. 다른 이름으로 저장:")
        print(f"     src/main/resources/excel-templates/chart-template.xlsx")
    except Exception as e:
        print(f"❌ 파일 저장 실패: {e}")

if __name__ == "__main__":
    print("=" * 60)
    print("📊 엑셀 기본 템플릿 생성")
    print("=" * 60)
    print()
    
    create_excel_template_simple()
    
    print()
    print("=" * 60)

